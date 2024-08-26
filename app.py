import streamlit as st
import folium
from streamlit_folium import st_folium
import pandas as pd
import openrouteservice
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font
from openpyxl.worksheet.hyperlink import Hyperlink
import os

from GAS import getOriginData

def get_duration(start,end):

    API_KEY = st.secrets["OPEN_ROUTE_API_KEY"]

    # 初始化OpenRouteService客户端
    client = openrouteservice.Client(key=API_KEY)

    routes = client.directions([start, end], profile='driving-car')
    if routes and 'routes' in routes and routes['routes']:
        route = routes['routes'][0]
        duration = route['summary']['duration'] / 60  # 转换为分钟
        return duration
    
def generate_itinerary(data, start_time_str):
    # 將開始時間字串轉換為 datetime 物件

    # print(data)

    start_time = datetime.strptime(start_time_str, '%H:%M')
    itinerary = []

    for i, entry in enumerate(data):

        # 如果不是第一個水路，先加入移動時間
        if i > 0:
            move_start_time = start_time
            move_end_time = move_start_time + timedelta(minutes=entry['移動時間'])
            itinerary.append([f"{move_start_time.strftime('%H:%M')}~{move_end_time.strftime('%H:%M')}", '路程'])
            start_time = move_end_time

        # 生成水路名稱的行程
        waterway_start_time = start_time
        waterway_end_time = waterway_start_time + timedelta(minutes=entry['停留時間'])
        itinerary.append([f"{waterway_start_time.strftime('%H:%M')}~{waterway_end_time.strftime('%H:%M')}", entry['水路名稱']])

        # 更新開始時間為水路停留結束時間
        start_time = waterway_end_time

    return itinerary

def get_coordinates(waterway_name):
    for route in st.session_state['routes']:
        if route["水路名稱"] == waterway_name:
            return route["經度"], route["緯度"]
    return None

def render_page1():

    col1,col2=st.columns([1,3])

    with col1:

        st.subheader("會勘基本資料")
        route_name = st.text_input('水路名稱')
        route_lon = st.text_input("經度")
        route_lat = st.text_input("緯度")

        if st.button("儲存位置"):
            # Validate that both latitude and longitude are provided
            if route_lat and route_lon:
                # Append new coordinates to the list
                st.session_state['coords'].append({'lat': route_lat, 'lng': route_lon})

                new_row = {
                    "序號": None,
                    "鄉鎮": None,
                    "水路名稱": route_name,
                    "工作站": None,
                    "水路長度": None,
                    "概估經費": 0,
                    "工程用地": None,
                    "水路用地": None,
                    "最佳施工期": None,
                    "經度": float(route_lat),
                    "緯度": float(route_lon),
                    "停留時間": 20.0,
                    "移動時間": None,
                    "計算時間": None
                }

                # Add the new row to the list
                st.session_state['routes'].append(new_row)
                st.success("位置已儲存!")
            else:
                st.error("請確保經度和緯度都已填寫。")
            
    with col2:

        st.subheader("會勘集合地點")

        # 定義地圖的初始位置和縮放級別
        initial_location = [23.7089, 120.5406]  # 這裡使用台中的經緯度
        initial_zoom = 10

        # 添加點擊事件處理
        def add_marker(folium_map, lat, lon, label):
            folium.Marker(
                location=[lat, lon],
                popup=f"{label}<br>經度: {lon}<br>緯度: {lat}",
                icon=folium.Icon(icon="info-sign")
            ).add_to(folium_map)

        # 創建一個 Folium 地圖
        map = folium.Map(location=initial_location, zoom_start=initial_zoom)

        # 顯示儲存的標記
        for i, coord in enumerate(st.session_state['coords']):
            add_marker(map, coord['lat'], coord['lng'], label=f"點 {i + 1}")

        if st.toggle("顯示紀錄", st.session_state['showMap']):

            df = pd.DataFrame(st.session_state.routes)

            for idx, row in df.iterrows():
                folium.Marker([row['緯度'], row['經度']], 
                            popup=f"{row['水路名稱']} ({row['經度']}, {row['緯度']})",
                            tooltip=row['水路名稱']).add_to(map)


        # 顯示 Folium 地圖並捕捉點擊事件
        map_data = st_folium(map, width=1000, height=500)

        # add_button=st.button('新增會勘位置',type='primary')

        # 如果有點擊事件，獲取點擊的位置
        if map_data and map_data['last_clicked']:
            lat = round(map_data['last_clicked']['lat'],6)
            lng = round(map_data['last_clicked']['lng'],6) # map_data['last_clicked']['lng']

            # 顯示暫存的坐標
            st.write(f"X座標: {lat}, Y座標: {lng}")

    # 顯示儲存按鈕

    # if add_button:
    #     if route_name=="":
    #         st.sidebar.warning("請輸入水路名稱")
    #     else:
    #     # if len(st.session_state['coords'])==0: 
    #         # st.session_state['coords'].append({'lat': lat, 'lng': lng})
    #         # new_row = {'序號': None, '水路名稱': route_name, '經度': lon, '緯度': lat, '停留時間': None, '移動時間': None}
            
    #         # new_row = {
    #         # "序號": None,
    #         # "水路名稱": route_name,
    #         # "經度": lng,
    #         # "緯度": lat,
    #         # "停留時間": None,
    #         # "移動時間": None
    #         # }

    #         new_row={
    #         "序號": None,
    #         "鄉鎮":None,
    #         "水路名稱": route_name,
    #         "工作站":None,
    #         "水路長度":None,
    #         "概估經費":0,
    #         "工程用地":None,
    #         "水路用地":None,
    #         "最佳施工期":None,
    #         "經度": st.session_state['coords']['lng'],
    #         "緯度": st.session_state['coords']['lat'],
    #         "停留時間": 20.0,
    #         "移動時間":None,
    #         "計算時間":None
    #         }

    #         # 将新的行添加到列表中
    #         st.session_state['routes'].append(new_row)
    #         st.session_state['showMap']=False
            
    #         # st.session_state['routes'] = st.session_state['routes'].append(new_row, ignore_index=True)
    #         st.rerun()
    #     # else:
    #         # st.write(len(st.session_state['coords'])==0)
    #         # st.sidebar.warning("會勘地點已經被安排")
                
    
    # st.json(st.session_state)

def render_page3():

    st.subheader("工程概要表")

    st.session_state.code=st.text_input("請輸入密碼",type="password")

    if st.session_state.code == st.secrets["CODE"]:

        df=getOriginData()

    # print(df)

    df_edit = st.data_editor(df,hide_index=True)

    df_meeting=df_edit[df_edit['meeting']]

    st.markdown("---")

    st.subheader("選取內容")

    df_result=[]

    for index,row in df_meeting.iterrows():
        st.write(f"ID: {row['id']}, 水路名稱: {row['inf.work_name']}, E: {row['coords.2.lon']}, N: {row['coords.2.lat']}")
        
        # 最佳施工期
        work_start_date_str = row['inf.work_start_date']
        work_end_date_str = row['inf.work_end_date']

        work_start_date=datetime.strptime(work_start_date_str, '%Y-%m-%d')
        work_end_date=datetime.strptime(work_end_date_str, '%Y-%m-%d')

        start_date_str = work_start_date.strftime("%Y/%m") if work_start_date else "未指定"
        end_date_str = work_end_date.strftime("%Y/%m") if work_end_date else "未指定"

        work_date_range= f"{start_date_str} ~ {end_date_str}"

        new_row={
            "序號": None,
            "鄉鎮":row['inf.work_place2'],
            "水路名稱": row['inf.work_name'],
            "工作站":row['inf.work_station'],
            "水路長度":row['inf.job_length'],
            "概估經費":row['inf.job_cost'],
            "工程用地":row['inf.work_place_water'],
            "水路用地":row['inf.work_place_detail'], 
            "最佳施工期":work_date_range,
            "經度": float(row['coords.2.lon']),
            "緯度": float(row['coords.2.lat']),
            "停留時間": 20.0,
            "移動時間":None,
            "計算時間":None
            }
        df_result.append(new_row)

    st.session_state['routes']= df_result

    # st.json(df_result)

def render_page2():

    st.write('#### 會勘清單')

    # 排序更新后的数据

    meet_date = st.date_input('會勘日期')
    meet_time = st.time_input('會勘時間')

    start_time_datetime = meet_time

    sorted_data=st.session_state['routes']

    # st.json(sorted_data)

    if any(item['序號'] is None for item in sorted_data):

        sorted_data = sorted(st.session_state['routes'], key=lambda x: x["經度"])

        for index, item in enumerate(sorted_data):
            item["序號"] = index+1

    # 使用 st.data_editor 允许用户编辑数据
    edited_data = st.data_editor(sorted_data)

    plot_data=sorted(edited_data, key=lambda x: x["序號"]) 

    if st.sidebar.button("計算路程時間"):

        df = pd.DataFrame(plot_data)

        # 提取经纬度并成对组合
        coordinates = df[['經度', '緯度']].values.tolist()
        paired_coordinates = [[coordinates[i], coordinates[i+1]] for i in range(len(coordinates) - 1)]

        for i, (start, end) in enumerate(paired_coordinates):
            api_time=get_duration(start,end)
            df.at[i+1,'計算時間'] = api_time
            # 使用 next 和生成器简化移動時間的赋值
            intervals = [5, 10, 15, 20, 25, 30]
            move_time = next((t for t in intervals if api_time < t), None)
            df.at[i + 1, '移動時間'] = move_time

        df.at[0,'移動時間'] = 0
        df.at[0,'計算時間'] = 0

        st.session_state['routes'] = df.to_dict('records')

        st.rerun()

    if st.sidebar.button("行程表"):

        start_time=start_time_datetime.strftime('%H:%M')

        my_list=(generate_itinerary(st.session_state['routes'],start_time))

        # 載入 Excel 檔案
        wb = load_workbook('./template/ARRANGE.xlsx')
        ws = wb.active
        ws.cell(row=2, column=4, value=meet_date.strftime("%Y-%m-%d"))

        # 取得活頁簿的最後一列
        last_row = 3
        j=0

        df_data=st.session_state['routes']

        for item in my_list:

            target_name = item[1]

            # 使用列表推导式筛选出符合条件的字典
            list_filter = [item for item in df_data if item.get('水路名稱') == target_name]

            if len(list_filter) > 0:

                dict_filter= list_filter[0]
                # print(dict_filter)
                # print(dict_filter['鄉鎮'])

                serial_number = item[0]
                waterway_name = item[1]
                coordinates = get_coordinates(waterway_name)

                if coordinates :
                    j=j+1
                    longitude, latitude = coordinates
                    url = f"https://www.google.com/maps/place/{latitude},{longitude}"

                    # 放置資料到相應的單元格
                    ws.cell(row=last_row + 1, column=1, value=j)
                    ws.cell(row=last_row + 1, column=3, value=dict_filter['鄉鎮'])
                    ws.cell(row=last_row + 1, column=4, value=waterway_name)
                    ws.cell(row=last_row + 1, column=5, value=dict_filter['工作站'])
                    ws.cell(row=last_row + 1, column=6, value=dict_filter['水路長度'])
                    ws.cell(row=last_row + 1, column=7, value=dict_filter['概估經費']/1000)
                    ws.cell(row=last_row + 1, column=8, value=dict_filter['水路用地'])
                    ws.cell(row=last_row + 1, column=9, value=dict_filter['工程用地'])
                    ws.cell(row=last_row + 1, column=10, value=dict_filter['最佳施工期'])
                    ws.cell(row=last_row + 1, column=11, value=serial_number)
                    ws.cell(row=last_row + 2, column=12, value=longitude)
                    ws.cell(row=last_row + 2, column=13, value=latitude)
                    cell=ws.cell(row=last_row + 1, column=12, value=url)
                    cell.hyperlink = url  # 設定超連結
                    cell.style = "Hyperlink"  # 設定超連結樣式
                    cell.font=Font(size=18, color="3399FF", underline="single")
                    # 合併儲存格
                    ws.merge_cells(start_row=last_row + 1, start_column=12, end_row=last_row + 1, end_column=13)
                    ws.cell(row=last_row + 1, column=12).alignment = Alignment(vertical='center', horizontal='center')
                    ws.cell(row=last_row + 1, column=12).font=Font(size=12, color="3399FF", underline="single")
                # else:

                #     ws.cell(row=last_row + 1, column=1, value=j)
                #     ws.cell(row=last_row + 1, column=4, value="...")
                #     ws.cell(row=last_row + 1, column=11, value=serial_number)
                #     ws.cell(row=last_row + 2, column=12, value="")
                #     ws.cell(row=last_row + 2, column=13, value="")
                #     ws.cell(row=last_row + 1, column=12, value="")
                #     # 合併儲存格
                #     ws.merge_cells(start_row=last_row + 1, start_column=12, end_row=last_row + 2, end_column=13)

                    last_row += 2  

            ws.print_area = 'A1:N'+str(last_row)


        # 儲存 Excel 檔案

        output_file = 'example.xlsx'

        wb.save(output_file)
        with open(output_file, 'rb') as f:
            bytes_data = f.read()
        st.sidebar.download_button(label='計算成果下載', data=bytes_data, file_name=output_file, type='primary')
        os.remove(output_file)


    ## 地圖部分
    st.write('#### 會勘地圖')

    IsShowPath = st.checkbox("是否顯示路徑?", True)

    map_data = folium.Map(location=[23.7089, 120.5406],zoom_start=11)  # 台湾的中心点

    if IsShowPath:
        path = [(item["緯度"], item["經度"]) for item in plot_data]
        folium.PolyLine(path, color="blue", weight=2.5, opacity=1).add_to(map_data)

        for index, item in enumerate(plot_data):
            # 添加带编号的标记
            folium.Marker(
                location=[item["緯度"], item["經度"]],
                icon=folium.DivIcon(
                    html=f"""
                        <div style="position: relative; top: -10px; left: -10px;">
                            <svg>
                                <circle cx="10" cy="10" r="10" fill="red" />
                                <text x="10" y="15" fill="white" text-anchor="middle" font-size="16">{item['序號']}</text>
                            </svg>
                        </div>
                    """
                )
            ).add_to(map_data)

        # map_path = 'map.html'
        # map_data.save(map_path)
    else:
        for item in sorted_data:
            folium.Marker([item['緯度'], item['經度']], 
                          popup=f"{item['水路名稱']} ({item['經度']}, {item['緯度']})",
                          tooltip=item['水路名稱'], icon=folium.Icon(color='green', prefix='glyphicon')).add_to(map_data)

    folium.raster_layers.WmsTileLayer(
        url='http://maps.nlsc.gov.tw/S_Maps/wms',  # 示例 WMS 服务 URL，请根据需要替换
        layers='TOWN',
        name='鄉鎮市',
        format='image/png',
        transparent=True,
        opacity=0.5,
        control=True
    ).add_to(map_data)

    folium.LayerControl().add_to(map_data)

        # 渲染地图并获取地图状态
    st_folium(map_data, width=800, height=400)

def main():

    SYSTEM_VERSION="V0.3.1"

    st.set_page_config(
        page_title="會勘地點安排"+SYSTEM_VERSION,
        page_icon="🌐",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    if 'current_page' not in st.session_state:
        st.session_state['current_page'] = 'page1'

    if 'code' not in st.session_state:
        st.session_state['code'] = ''

    # 用來暫存點擊的座標
    if 'coords' not in st.session_state:
        st.session_state['coords'] = []

    if 'showMap' not in st.session_state:
        st.session_state['showMap'] = True

    # initial_data = [
    #     { "序號": None,"水路名稱": "北月眉小排2-1", "經度": 120.286, "緯度": 23.6901, "停留時間": 20.0,"移動時間":None,"計算時間":None},
    #     { "序號": None,"水路名稱": "路利潭中排一", "經度": 120.2368, "緯度": 23.7008, "停留時間": 20.0,"移動時間":None,"計算時間":None},
    #     { "序號": None,"水路名稱": "丘厝小排2-4", "經度": 120.2037, "緯度": 23.6724, "停留時間": 20.0,"移動時間":None,"計算時間":None},
    #     { "序號": None,"水路名稱": "牛厝中排1", "經度": 120.2066, "緯度": 23.6667, "停留時間": 20.0,"移動時間":None,"計算時間":None},
    # ]


    if 'routes' not in st.session_state:
        st.session_state['routes'] = [] #initial_data
        # st.session_state['routes'] = initial_data

    with st.sidebar:

        # st.json(st.session_state)
        st.title(":globe_with_meridians: 會勘地點安排 "+SYSTEM_VERSION)
        st.write("這是用於提報計畫時的估算工具")
        st.info("作者:**林宗漢**")


        st.markdown("---")

        st.subheader("選擇頁面")
        # if st.button("新增會勘地點"):
            # st.session_state.current_page = 'page1'  

        # st.session_state.code=st.text_input("請輸入密碼",type="password")

        # if st.session_state.code == st.secrets["CODE"]:

        # st.markdown("---")

        # st.subheader("選擇頁面")
        if st.button("新增會勘地點"):
            st.session_state.current_page = 'page1'  
        if st.button("安排會勘地點"):
            st.session_state.current_page = 'page2'
        if st.button("工程概要表"):
            st.session_state.current_page = 'page3'

        st.markdown("---")
        st.subheader("操作按鈕")


    if st.session_state.current_page == 'page1':
        render_page1()
    elif st.session_state.current_page == 'page2':
        render_page2()
    elif st.session_state.current_page == 'page3':
        render_page3()


    # print(st.session_state)

if __name__ == "__main__":
    main()